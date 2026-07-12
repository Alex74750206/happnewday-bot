import os
import glob
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

_DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
_LOG_FILE = os.path.join(_DATA_DIR, "songs_log.xlsx")


def get_log_file_path() -> str:
    """Путь к файлу лога заказов — используется чтобы отправить его админу в Telegram."""
    return _LOG_FILE

_HEADERS = ["Дата", "Время", "ID пользователя", "Имя в Telegram",
            "Имя получателя", "Повод", "Стиль", "Ссылка на песню", "Текст песни"]

# Ширина столбцов (символы)
_COL_WIDTHS = [12, 7, 18, 18, 18, 20, 20, 45, 60]
# Высота строк данных (пунктов)
_ROW_HEIGHT = 18

_HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
_LINK_FONT     = Font(color="0563C1", underline="single", size=10)
_DATA_FONT     = Font(size=10)
_ALIGN_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=False)
_ALIGN_LEFT    = Alignment(horizontal="left", vertical="center", wrap_text=False)
_ALIGN_LYRICS  = Alignment(horizontal="left", vertical="top", wrap_text=True)

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _new_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы"

    # Заголовки
    for col, (header, width) in enumerate(zip(_HEADERS, _COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    return wb


def _apply_row(ws, row_idx: int, values: list, audio_url: str) -> None:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col)

        if col == 8 and audio_url:               # Ссылка на песню
            cell.value = "▶ Слушать"
            cell.hyperlink = audio_url
            cell.font = _LINK_FONT
            cell.alignment = _ALIGN_CENTER
        elif col == 9:                            # Текст песни
            cell.value = value
            cell.font = _DATA_FONT
            cell.alignment = _ALIGN_LYRICS
        elif col in (1, 2, 3):                   # Дата, Время, ID
            cell.value = value
            cell.font = _DATA_FONT
            cell.alignment = _ALIGN_CENTER
        else:
            cell.value = value
            cell.font = _DATA_FONT
            cell.alignment = _ALIGN_LEFT

        cell.border = _BORDER

    ws.row_dimensions[row_idx].height = _ROW_HEIGHT


def _merge_pending_tmp_files(ws) -> None:
    """Подхватывает записи из временных файлов (созданных, пока основной файл был занят Excel)
    и дописывает их в основной файл, чтобы они не терялись навсегда."""
    pattern = os.path.join(_DATA_DIR, "songs_log_tmp_*.xlsx")
    for tmp_path in sorted(glob.glob(pattern)):
        try:
            tmp_wb = load_workbook(tmp_path)
            tmp_ws = tmp_wb.active
            for r in range(2, tmp_ws.max_row + 1):
                if tmp_ws.cell(row=r, column=1).value is None:
                    continue
                values = [tmp_ws.cell(row=r, column=c).value for c in range(1, 8)]
                values.append("")  # placeholder для ссылки — сама ссылка ниже, через hyperlink
                values.append(tmp_ws.cell(row=r, column=9).value)
                link_cell = tmp_ws.cell(row=r, column=8)
                audio_url = link_cell.hyperlink.target if link_cell.hyperlink else ""
                _apply_row(ws, ws.max_row + 1, values, audio_url)
            os.remove(tmp_path)
        except Exception:
            continue  # повреждённый/всё ещё занятый tmp — попробуем подхватить в следующий раз


def log_song(user_id: int, username: str, name: str, occasion: str,
             style: str, audio_url: str = "", lyrics: str = "") -> None:
    os.makedirs(_DATA_DIR, exist_ok=True)

    now = datetime.now()
    values = [
        now.strftime("%d.%m.%Y"),
        now.strftime("%H:%M"),
        user_id,
        username or "—",
        name,
        occasion,
        style,
        "",        # placeholder для ссылки (заполняется отдельно)
        lyrics,
    ]

    try:
        if os.path.exists(_LOG_FILE):
            wb = load_workbook(_LOG_FILE)
            ws = wb.active
        else:
            wb = _new_workbook()
            ws = wb.active

        _merge_pending_tmp_files(ws)  # сначала подтягиваем всё отложенное, потом пишем новую строку

        row_idx = ws.max_row + 1
        _apply_row(ws, row_idx, values, audio_url)
        wb.save(_LOG_FILE)
    except Exception:
        # Файл занят (Excel открыт) или иная ошибка записи — сохраняем ТОЛЬКО эту одну запись
        # в отдельный минимальный временный файл (не копию всего лога!), чтобы при следующем
        # успешном log_song() её можно было однозначно, без дублей, подхватить обратно.
        import time
        tmp_wb = _new_workbook()
        tmp_ws = tmp_wb.active
        _apply_row(tmp_ws, 2, values, audio_url)
        tmp = _LOG_FILE.replace(".xlsx", f"_tmp_{int(time.time())}.xlsx")
        tmp_wb.save(tmp)
